from openpyxl import load_workbook


def get_departments(excel_file: str, column: str) -> list:
    """
    Возвращает уникальные значений в столбце из файла. Предназначено для отедлов и управлений.
    :param excel_file: Путь к файлу Excel
    :param column: Столбец из которого необходимо считать уникальные значения (отдел/управления)
    :return values_list: Список уникальных значений в столбце
    """
    try:
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb['Лист1']
        target_column = column
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        column_index = headers.index(target_column)
        values = [str(row[column_index]) for row in sheet.iter_rows(min_row=2, values_only=True)]
        values_list = sorted(list(set(value for value in values if value not in ('None', ''))))
        return values_list
    except Exception as e:
        raise ValueError(f"Ошибка при извлечении отделов/управлений: {str(e)}")


def get_employees_by_department(excel_file: str, department: str) -> list:
    """
    Возвращает список сотрудников указанного отдела.
    :param excel_file: Путь к файлу Excel
    :param department: Название отдела
    :return filtered_rows: Список сотрудников
    """
    try:
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb['Лист1']
        department_index = 4
        unique_employees = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[department_index] == department:
                key = row[5]
                unique_employees[key] = [row[5], row[2], str(row[4])]
        return list(unique_employees.values())
    except Exception as e:
        raise ValueError(f"Ошибка при извлечении сотрудников: {str(e)}")


def get_employee_by_fio_department(excel_file: str, department: str, name: str):
    """
    Чтение информации о сотруднике по ФИО и отделу.
    :param excel_file: Путь к файлу для чтения.
    :param department: Название отдела.
    :param name: ФИО сотрудника.
    :return employees[name]:
        Информация о сотруднике в формате:
             {
                 "ФИО": str,
                 "Должность": str,
                 "Отдел": str,
                 "Результаты": [
                     {
                         "Год тестирования": str,
                         "Критерии": {"Критерий1": float, "Критерий2": float, ...},
                         "Сумма баллов": float
                     },
                     ...
                 ]
             }
    """
    try:
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb['Лист1']
        department_index = 4
        name_index = 5
        employees = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row[department_index] == department and row[name_index] == name)\
                    or (row[department_index] in ('', None) and row[name_index] == name):
                criteria = dict(zip(
                    [str(i) for i in range(1, 12)],
                    [float(i.replace(',', '.')) for i in row[6:17]]
                ))
                results = dict(zip(
                    ['Год тестирования', 'Критерии', 'Cумма баллов'],
                    [str(row[18]), criteria, float(row[17].replace(',', '.'))]
                ))
                fio = row[5]
                if fio not in employees:
                    employees[fio] = {
                        'ФИО': fio,
                        'Должность': row[2],
                        'Отдел': str(row[4]),
                        'Результаты': [results]
                    }
                else:
                    employees[fio]['Результаты'].append(results)
        return employees[name]
    except Exception as e:
        print(f"Ошибка при чтении данных: {str(e)}")
        return None


