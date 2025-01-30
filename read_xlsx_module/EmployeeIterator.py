from openpyxl import load_workbook


class GetEmployee:
    def __init__(self, excel_file: str, department=None, adminisrtation=None):
        """
        Инициализация итератора для сотрудников выбранного отдела и управления.
        Без аргументов кроме файла будут выдаваться все сотрудники института.
        :param excel_file: Путь к файлу для чтения
        :param department:str Название отдела
        :param adminisrtation:int Номер управления
        """
        try:
            self.wb = load_workbook(excel_file, data_only=True)
            self.sheet = self.wb['Лист1']
            self.department = department
            self.administration = adminisrtation
            self.rows = self._filter_rows()
            self.index = 0
        except Exception as e:
            raise ValueError(f"Ошибка при чтении файла: {str(e)}")

    def _filter_rows(self):
        """
        Фильтрует строки согласно выбраному отделу и/или управлению
        :return: Список словарей по шаблону как в функции get_employee_by_fio_department
        """
        department_index = 4
        admin_index = 0
        employees = {}
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if (self.department is None or row[department_index] == self.department)\
                    and (self.administration is None or row[admin_index] == self.administration):
                criteria = dict(zip(
                    [str(i) for i in range(1, 12)],
                    [float(j.replace(',', '.')) for j in row[6:17] if j is not None]
                ))
                results = dict(zip(
                    ['Год тестирования', 'Критерии', 'Cумма баллов'],
                    [str(row[18]), criteria, row[17]]
                ))
                fio = row[5]
                if fio not in employees:
                    employees[fio] = {
                        'ФИО': fio,
                        'Должность': row[2],
                        'Отдел': str(row[4]),
                        'Результаты': [results]
                    }
                else:
                    employees[fio]['Результаты'].append(results)
        return list(employees.values())

    def __iter__(self):
        return self

    def __next__(self):
        """
        Возвращает данные об одном сотруднике за итерацию
        """
        if self.index < len(self.rows):
            current_worker = self.rows[self.index]
            self.index += 1
            return current_worker
        else:
            self.wb.close()
            raise StopIteration
